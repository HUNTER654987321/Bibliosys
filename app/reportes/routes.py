from datetime import date
from io import BytesIO
from flask import Blueprint, render_template, send_file
from flask_login import login_required
from sqlalchemy import func, desc
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from app import db
from app.decorators import rol_requerido
from app.models import Usuario, Libro, Prestamo, Multa, Categoria, obtener_configuracion_sistema

reportes_bp = Blueprint("reportes", __name__, url_prefix="/reportes")

def _datos_generales():
    prestamos = Prestamo.query.all()
    return {
        "usuarios": Usuario.query.count(),
        "libros": Libro.query.count(),
        "prestamos": Prestamo.query.count(),
        "activos": sum(1 for p in prestamos if p.estado_actual == "Activo"),
        "devueltos": sum(1 for p in prestamos if p.estado_actual == "Devuelto"),
        "vencidos": sum(1 for p in prestamos if p.estado_actual == "Vencido"),
        "multas": Multa.query.filter_by(pagado=False).count(),
    }

def _libros_populares(limit=None):
    query = db.session.query(
        Libro,
        func.count(Prestamo.id).label("total_prestamos")
    ).join(Prestamo).group_by(Libro.id).order_by(desc("total_prestamos"), Libro.titulo.asc())
    if limit:
        query = query.limit(limit)
    return query.all()

@reportes_bp.route("/")
@login_required
@rol_requerido("Administrador", "Bibliotecario")
def index():
    datos = _datos_generales()
    categorias = Categoria.query.all()
    prestamos = Prestamo.query.order_by(Prestamo.id.asc()).limit(10).all()
    multas = Multa.query.order_by(Multa.id.asc()).limit(10).all()
    populares = _libros_populares(limit=10)
    return render_template(
        "reportes/index.html",
        datos=datos,
        categorias=categorias,
        prestamos=prestamos,
        multas=multas,
        populares=populares,
        today=date.today()
    )

def crear_pdf(titulo, encabezados, filas):
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet

    config = obtener_configuracion_sistema()
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), leftMargin=28, rightMargin=28, topMargin=28, bottomMargin=28)
    styles = getSampleStyleSheet()
    elementos = [
        Paragraph(config.nombre_biblioteca or "BiblioSys", styles['Heading2']),
        Paragraph(titulo, styles['Title']),
        Paragraph(f"Fecha: {date.today()}", styles['Normal']),
        Spacer(1, 14)
    ]
    data = [encabezados] + filas
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f766e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dce5e0')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f6faf8')]),
    ]))
    elementos.append(table)
    doc.build(elementos)
    buffer.seek(0)
    return buffer

def crear_excel(titulo, encabezados, filas):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"
    ws.append([titulo])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(encabezados), 1))
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.append([f"Fecha: {date.today()}"])
    ws.append([])
    ws.append(encabezados)

    fill = PatternFill("solid", fgColor="0F766E")
    for cell in ws[4]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

    for fila in filas:
        ws.append(fila)

    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            value = str(value) if value is not None else ""
            max_length = max(max_length, len(value))
        ws.column_dimensions[column].width = min(max_length + 3, 45)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def _filas_libros():
    libros = Libro.query.order_by(Libro.titulo.asc()).all()
    return [[
        l.id,
        l.titulo,
        ", ".join(a.nombre for a in l.autores) or "Sin autor",
        l.isbn,
        l.categoria.nombre,
        l.editorial.nombre,
        l.stock,
        l.disponible
    ] for l in libros]

def _filas_prestamos():
    prestamos = Prestamo.query.order_by(Prestamo.id.asc()).all()
    return [[
        p.id,
        p.usuario.nombre,
        p.libro.titulo,
        str(p.fecha_prestamo),
        str(p.fecha_devolucion),
        p.estado_actual,
        f"Bs {p.multa.monto:.2f}" if p.multa else "Sin multa"
    ] for p in prestamos]

def _filas_usuarios():
    usuarios = Usuario.query.order_by(Usuario.id.asc()).all()
    return [[u.id, u.nombre, u.email, u.rol.nombre, 'Activo' if u.activo else 'Inactivo', len(u.prestamos)] for u in usuarios]

def _filas_populares():
    return [[libro.id, libro.titulo, ", ".join(a.nombre for a in libro.autores) or "Sin autor", total, libro.stock, libro.disponible] for libro, total in _libros_populares()]

@reportes_bp.route('/libros/pdf')
@login_required
@rol_requerido("Administrador", "Bibliotecario")
def libros_pdf():
    pdf = crear_pdf('Reporte de libros', ['ID','Titulo','Autor(es)','ISBN','Categoria','Editorial','Stock','Disp.'], _filas_libros())
    return send_file(pdf, as_attachment=True, download_name='reporte_libros.pdf', mimetype='application/pdf')

@reportes_bp.route('/libros/excel')
@login_required
@rol_requerido("Administrador", "Bibliotecario")
def libros_excel():
    excel = crear_excel('Reporte de libros', ['ID','Titulo','Autor(es)','ISBN','Categoria','Editorial','Stock','Disp.'], _filas_libros())
    return send_file(excel, as_attachment=True, download_name='reporte_libros.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@reportes_bp.route('/prestamos/pdf')
@login_required
@rol_requerido("Administrador", "Bibliotecario")
def prestamos_pdf():
    pdf = crear_pdf('Reporte de prestamos', ['ID','Usuario','Libro','Prestamo','Entrega','Estado','Multa'], _filas_prestamos())
    return send_file(pdf, as_attachment=True, download_name='reporte_prestamos.pdf', mimetype='application/pdf')

@reportes_bp.route('/prestamos/excel')
@login_required
@rol_requerido("Administrador", "Bibliotecario")
def prestamos_excel():
    excel = crear_excel('Reporte de prestamos', ['ID','Usuario','Libro','Prestamo','Entrega','Estado','Multa'], _filas_prestamos())
    return send_file(excel, as_attachment=True, download_name='reporte_prestamos.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@reportes_bp.route('/usuarios/pdf')
@login_required
@rol_requerido("Administrador", "Bibliotecario")
def usuarios_pdf():
    pdf = crear_pdf('Reporte de usuarios', ['ID','Nombre','Correo','Rol','Estado','Prestamos'], _filas_usuarios())
    return send_file(pdf, as_attachment=True, download_name='reporte_usuarios.pdf', mimetype='application/pdf')

@reportes_bp.route('/usuarios/excel')
@login_required
@rol_requerido("Administrador", "Bibliotecario")
def usuarios_excel():
    excel = crear_excel('Reporte de usuarios', ['ID','Nombre','Correo','Rol','Estado','Prestamos'], _filas_usuarios())
    return send_file(excel, as_attachment=True, download_name='reporte_usuarios.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@reportes_bp.route('/populares/pdf')
@login_required
@rol_requerido("Administrador", "Bibliotecario")
def populares_pdf():
    pdf = crear_pdf('Reporte de libros mas prestados', ['ID','Titulo','Autor(es)','Prestamos','Stock','Disponible'], _filas_populares())
    return send_file(pdf, as_attachment=True, download_name='reporte_libros_populares.pdf', mimetype='application/pdf')

@reportes_bp.route('/populares/excel')
@login_required
@rol_requerido("Administrador", "Bibliotecario")
def populares_excel():
    excel = crear_excel('Reporte de libros mas prestados', ['ID','Titulo','Autor(es)','Prestamos','Stock','Disponible'], _filas_populares())
    return send_file(excel, as_attachment=True, download_name='reporte_libros_populares.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
